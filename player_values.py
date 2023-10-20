import os
import pandas as pd
import requests
from openpyxl import load_workbook
from io import StringIO

# GitHub CSV URL
csv_url = 'https://raw.githubusercontent.com/dynastyprocess/data/master/files/values.csv'

# Excel file details
excel_file = 'player_values.xlsx'
sheet_name = 'Player Values'

def update_excel_with_csv_data(csv_url, excel_file, sheet_name):
    # Load data from the CSV file
    csv_data = requests.get(csv_url).text
    df = pd.read_csv(StringIO(csv_data))

    # Check if the Excel file exists, if not, create a new one
    if not os.path.exists(excel_file):
        df.to_excel(excel_file, sheet_name=sheet_name, index=False)
        print(f"New Excel file '{excel_file}' created with data from {csv_url}.")
    else:
        # Load existing Excel workbook
        workbook = load_workbook(excel_file)

        # Check if the sheet already exists, if not, create one
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)

        # Select the sheet and write data from the DataFrame
        writer = pd.ExcelWriter(excel_file, engine='openpyxl')
        writer.book = workbook
        writer.sheets = {ws.title: ws for ws in workbook.worksheets}
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Save the changes
        writer.save()
        writer.close()

        print(f"Data from {csv_url} successfully updated in {excel_file} ({sheet_name}).")

if __name__ == "__main__":
    update_excel_with_csv_data(csv_url, excel_file, sheet_name)
